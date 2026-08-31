import os
import re
import openpyxl
from typing import Dict, List, Any, Optional

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATASET_PATH = os.path.join(BASE_DIR, "assest", "excle", "MCOC_dataset.xlsx")
IMAGES_DIR = os.path.join(BASE_DIR, "frontend", "public", "images")
CLASSES_DIR = os.path.join(IMAGES_DIR, "classes")

def normalize_key(s: str) -> str:
    if not s:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

import json

IMAGE_MAP: Dict[str, str] = {}
IMAGE_MAP_PATH = os.path.join(BASE_DIR, "backend", "image_map.json")
if os.path.exists(IMAGE_MAP_PATH):
    try:
        with open(IMAGE_MAP_PATH, "r") as f:
            IMAGE_MAP = json.load(f)
    except Exception as e:
        print(f"[!] Error loading image_map.json: {e}")

# Comprehensive explicit champion name aliases
# Covers MCOC acronyms, abbreviations, distinct champion variants, and user-renamed files
ALIASES = {
    # The Champion & Okoye & Classic Spidey (Newly Renamed by User)
    normalize_key("The Champion"): "The-champion.png",
    normalize_key("The-Champion"): "The-champion.png",
    normalize_key("Champion"): "The-champion.png",
    normalize_key("Okoye"): "okoye.png",
    normalize_key("Spider-Man (Classic)"): "spiderman(Classic).png",
    normalize_key("Spider-Man (Classic)"): "spiderman(Classic).png",
    normalize_key("Spider all"): "spiderman(Classic).png",
    normalize_key("Spider-Man"): "spiderman(Classic).png",
    normalize_key("Classic Spider-Man"): "spiderman(Classic).png",

    # Magneto variants
    normalize_key("Magneto"): "magneto.png",
    normalize_key("Magneto (House of X)"): "Magneto (House of X).png",
    normalize_key("Magneto (White)"): "Magneto (House of X).png",
    normalize_key("Magneto(white)"): "Magneto (House of X).png",
    normalize_key("White Magneto"): "Magneto (House of X).png",
    normalize_key("Magneto House of X"): "Magneto (House of X).png",
    
    # Iron Man variants
    normalize_key("Iron Man"): "ironman.webp",
    normalize_key("Iron Man (Infamous)"): "Iron Man (Infamous).png",
    normalize_key("Infamous Iron man"): "Iron Man (Infamous).png",
    normalize_key("Infamous Iron Man"): "Iron Man (Infamous).png",
    normalize_key("Iron Man (Infinity War)"): "Iron Man (Infinity War).png",
    normalize_key("Superior Iron Man"): "superiorironman.png",
    normalize_key("Iron Patriot"): "ironpatriot.png",
    normalize_key("Ironheart"): "Ironheart.webp",
    normalize_key("IronHeart"): "Ironheart.webp",

    # Hulk variants
    normalize_key("Hulk"): "hulk.png",
    normalize_key("Hulk (Immortal)"): "hulkimmortal.png",
    normalize_key("Hulk (imortal)"): "hulkimmortal.png",
    normalize_key("Immortal Hulk"): "hulkimmortal.png",
    normalize_key("Hulk (Ragnarok)"): "hulkragnarok.png",
    normalize_key("Gladiator Hulk"): "hulkragnarok.png",
    normalize_key("Hulkbuster"): "hulkbuster.png",
    normalize_key("Hulkling"): "hulkling.png",
    normalize_key("Hulking"): "hulkling.png",
    normalize_key("Red Hulk"): "redhulk.png",
    normalize_key("Joe Fixit"): "joefixit.png",
    normalize_key("Jeo Fixit"): "joefixit.png",

    # Spider-Man variants
    normalize_key("Spider-Man (Miles Morales)"): "spidermanmilesmorales.png",
    normalize_key("Miles Morales"): "spidermanmilesmorales.png",
    normalize_key("miles"): "spidermanmilesmorales.png",
    normalize_key("Milse"): "spidermanmilesmorales.png",
    normalize_key("Spider-Man (Stealth Suit)"): "spidermanstealthsuit.png",
    normalize_key("Stealth Suit Spider-Man"): "spidermanstealthsuit.png",
    normalize_key("Spider-Man (Stark Enhanced)"): "spidermanstarkenhanced.png",
    normalize_key("Stark Enhanced Spider-Man"): "spidermanstarkenhanced.png",
    normalize_key("Spider-Man (Symbiote)"): "spidermansymbiote.webp",
    normalize_key("Spider-Man (Supreme)"): "spiderman_supreme.png",
    normalize_key("Spider-Man (Pavitr Prabhakar)"): "spiderman_pavitr.png",
    normalize_key("Spider-Man (Pavitra Prabhakar)"): "spiderman_pavitr.png",
    normalize_key("SPIDER MAN ( PAVITRA PRABHAKAR )"): "spiderman_pavitr.png",
    normalize_key("Spider Man ( Pavitra Prabhakar )"): "spiderman_pavitr.png",
    normalize_key("Pavitra prabakar"): "spiderman_pavitr.png",
    normalize_key("Spider-Man 2099"): "spiderman2099.png",
    normalize_key("Spider-Ham"): "spiderham.png",
    normalize_key("Spider-Punk"): "spider-punk.webp",
    normalize_key("Spider-Woman"): "Spider-Woman.webp",
    normalize_key("SPIDER-WOMAN"): "Spider-Woman.webp",
    normalize_key("Spider-Woman (Jessica Drew)"): "Spider-Woman.webp",
    normalize_key("SPIDER-WOMAN (JESSICA DREW)"): "Spider-Woman.webp",
    normalize_key("Spider-Gwen"): "spidergwen.png",
    normalize_key("Spider-Slayer"): "Spider-Slayer_jjj_portrait.webp",
    normalize_key("Spider-Slayer (J. Jonah Jameson)"): "Spider-Slayer_jjj_portrait.webp",

    # Jean Grey / Jean Gray variants
    normalize_key("Jean Grey"): "jean_grey.webp",
    normalize_key("JEAN GREY"): "jean_grey.webp",
    normalize_key("Jean Gray"): "jean_grey.webp",
    normalize_key("JEAN GRAY"): "jean_grey.webp",

    # Daredevil Hell's Kitchen variants
    normalize_key("Daredevil (Hell's Kitchen)"): "Daredevil (Hell's Kitchen).png",
    normalize_key("DAREDEVIL (HELL'S KITCHEN)"): "Daredevil (Hell's Kitchen).png",
    normalize_key("Daredevil(H.K.)"): "Daredevil (Hell's Kitchen).png",
    normalize_key("DAREDEVIL(H.K.)"): "Daredevil (Hell's Kitchen).png",
    normalize_key("Daredevil (H.K.)"): "Daredevil (Hell's Kitchen).png",
    normalize_key("Daredevil HK"): "Daredevil (Hell's Kitchen).png",

    # Captain America Civil War / Civil Warrior
    normalize_key("CAPTAIN AMERICA (CIVIL WAR)"): "civilwarrior.png",
    normalize_key("Captain America (Civil War)"): "civilwarrior.png",
    normalize_key("Civil Warrior"): "civilwarrior.png",
    normalize_key("Civil Warrior (Captain America)"): "civilwarrior.png",

    # Wolverine variants
    normalize_key("Wolverine"): "wolverine.png",
    normalize_key("Wolverine (Weapon X)"): "wolverineweaponx.png",
    normalize_key("Weapon X"): "wolverineweaponx.png",
    normalize_key("Wepon X"): "wolverineweaponx.png",
    normalize_key("Old Man Logan"): "oldmanlogan.png",
    normalize_key("Wolverine (X 23)"): "x23.png",
    normalize_key("Wolverine (X-23)"): "x23.png",
    normalize_key("X-23"): "x23.png",
    normalize_key("X23"): "x23.png",

    # Ghost Rider & CGR
    normalize_key("Ghost Rider"): "ghostrider.webp",
    normalize_key("Cosmic Ghost Rider"): "cosmicghostrider.png",
    normalize_key("CGR"): "cosmicghostrider.png",

    # Black Widow & Claire Voyant
    normalize_key("Black Widow"): "blackwidow.png",
    normalize_key("Black Widow (Claire Voyant)"): "blackwidowvlairevoyant.png",
    normalize_key("Black Widow (Deadly Origin)"): "blackwidowwhite.png",
    normalize_key("BLCV"): "blackwidowvlairevoyant.png",

    # Doctor Doom & Doctor Strange & Voodoo
    normalize_key("Doctor Doom"): "doctordoom.png",
    normalize_key("Doom"): "doctordoom.png",
    normalize_key("Dr Doom"): "doctordoom.png",
    normalize_key("Doctor Strange"): "Doctor_Strange_portrait.webp",
    normalize_key("Doctor Voodoo"): "doctorvoodoo.png",
    normalize_key("VOODOO"): "doctorvoodoo.png",
    normalize_key("Doctor Bong"): "Doctor_Bong_portrait.webp",

    # Serpent & Hercules & Onslaught
    normalize_key("The Serpent"): "the_serpent.webp",
    normalize_key("Serpent"): "the_serpent.webp",
    normalize_key("Hercules"): "hercules.png",
    normalize_key("Onslaught"): "onslaughtt.webp",

    # Absorbing Man & Archangel & Apocalypse
    normalize_key("Absorbing Man"): "Absorbing_Man.png",
    normalize_key("ABS man"): "Absorbing_Man.png",
    normalize_key("Archangel"): "archangel.png",
    normalize_key("Archangle"): "archangel.png",
    normalize_key("Apocalypse"): "apocalypse.png",
    normalize_key("Apocalypes"): "apocalypse.png",

    # Other acronyms & Catalog Fallbacks
    normalize_key("Nico Minoru"): "nicominoru.png",
    normalize_key("Nicro"): "nicominoru.png",
    normalize_key("Hit-Monkey"): "hitmonkey.png",
    normalize_key("Hit Mokey"): "hitmonkey.png",
    normalize_key("Nick Fury"): "nickfury.png",
    normalize_key("NickFury"): "nickfury.png",
    normalize_key("Human Torch"): "humantorch.png",
    normalize_key("Human Tourch"): "humantorch.png",
    normalize_key("Nimrod"): "nimrod.png",
    normalize_key("Nimord"): "nimrod.png",
    normalize_key("Shuri"): "shuri.png",
    normalize_key("shuri"): "shuri.png",
    normalize_key("Viv Vision"): "Viv_Vision.png",
    normalize_key("VIV Vision"): "Viv_Vision.png",
    normalize_key("Chee'ilth"): "chee'ilth.webp",
    normalize_key("Cheelith"): "chee'ilth.webp",
    normalize_key("Baron Zemo"): "Baron_Zemo.png",
    normalize_key("Baran Zemo"): "Baron_Zemo.png",
    normalize_key("Hawkeye"): "hawkeye.png",
    normalize_key("Hawkey"): "hawkeye.png",
    normalize_key("Rocket Raccoon"): "Rocket Raccoon.png",
    normalize_key("Rocker Raccon"): "Rocket Raccoon.png",
    normalize_key("Captain Britain"): "captainbritain.png",
    normalize_key("Capatain Britain"): "captainbritain.png",
    normalize_key("Mr. Knight"): "Mr. Knight.png",
    normalize_key("MR.knight"): "Mr. Knight.png",
    normalize_key("Guillotine 2099"): "guillotine2099.png",
    normalize_key("G 2099"): "guillotine2099.png",
    normalize_key("Count Nefaria"): "count_nefaria.png",
    normalize_key("Count nefaria"): "count_nefaria.png",
    normalize_key("Warlock"): "warlock.png",
    normalize_key("Adam Warlock"): "adamwarlock.png",
    normalize_key("Aegon"): "aegon.png",
    normalize_key("Ægon"): "aegon.png",
    normalize_key("gon"): "aegon.png",
    normalize_key("Yondu"): "yondu.png",
    normalize_key("yondu"): "yondu.png",
    normalize_key("Longshot"): "longshot.png",
    normalize_key("LongShot"): "longshot.png",
    normalize_key("Anti-Venom"): "antivenom.png",
    normalize_key("Anti Venom"): "antivenom.png",
    normalize_key("She-Hulk"): "shehulk.webp",
    normalize_key("Shehulk"): "shehulk.webp",
    normalize_key("Mister Sinister"): "mistersinister.png",
    normalize_key("Mole Man"): "moleman.png",
    normalize_key("Mole man"): "moleman.png",
    normalize_key("Yelena Belova"): "portrait_yelenabelova.webp",
    normalize_key("Red Skull"): "redskull.png",
    normalize_key("Red Goblin"): "redgoblin.png",
    normalize_key("Beta Ray Bill"): "Beta_Ray_Bill.webp",
    normalize_key("Negasonic Teenage Warhead"): "Negasonic Teenage Warhead.webp",
    normalize_key("Star-Lord"): "starlord.png",
    normalize_key("Star-Lord (Stellar Forged)"): "Star-Lord2.webp",
    normalize_key("Madelyne Pryor"): "Madelyne_Pryor_portrait.webp",
    normalize_key("Ruby Thursday"): "Ruby_Thursday_portrait.webp",
    normalize_key("Thor (Ragnarok)"): "Thor_(Ragnarok)_portrait.webp",
    normalize_key("Thor (Jane Foster)"): "thorjanefoster.png",
    normalize_key("Wave"): "Wave_portrait.webp",
    normalize_key("Imperiosa"): "Imperiosa_portrait.webp",
    normalize_key("Pixie"): "Pixie_portrait.webp",
    normalize_key("Black Tarantula"): "blacktarantula.webp",
    normalize_key("Agatha Harkness"): "agathaharkness.webp",
    normalize_key("Bastion"): "bastion.webp",
    normalize_key("Arnim Zola"): "arimzola.webp",
    normalize_key("The Leader"): "The_ledger.webp",
    normalize_key("Leader"): "The_ledger.webp",
    normalize_key("Future Ant-Man"): "ant_man_future.webp",
    normalize_key("Ant-Man (Future)"): "ant_man_future.webp",
    normalize_key("Gentle"): "gentle_og.png",
    normalize_key("Dark Phoenix"): "phoenix_dark.webp",
    normalize_key("Phoenix"): "phoenix.webp",
    normalize_key("Silver Surfer"): "siversurfer.png",
    normalize_key("Ms. Marvel (Kamala Khan)"): "mskamalakhan.png",
    normalize_key("Captain Marvel (Classic)"): "captainmarvel.png",
    normalize_key("Captain Marvel"): "captainmarvelmovie.png",
    normalize_key("Blade (Stellar Forged)"): "blade2.png",
    normalize_key("Blade"): "blade.png",
    normalize_key("Isophyne"): "isophyne_latest.webp",
    normalize_key("Enchantress"): "enchantresss.webp",
    normalize_key("Scarlet Witch (Sigil)"): "scarletwitchmarvel.png",
    normalize_key("Scarlet Witch (Classic)"): "scarletwitch.png",
    normalize_key("Scarlet Witch"): "scarletwitch.png",
    normalize_key("The Hood"): "hood.png",
    normalize_key("Hood"): "hood.png",
    normalize_key("Blue Marvel"): "blue_marvel_portrait.webp",
    normalize_key("High Evolutionary"): "portrait_high_evolutionary.png",
    normalize_key("Jessica Jones"): "jessica.png",
    normalize_key("Captain America (Infinity War)"): "captainamerica.png",
    normalize_key("Captain America (Sam Wilson)"): "captainamericasw.png",
    normalize_key("Captain America (WWII)"): "captainamericawwii.png",
    normalize_key("Captain America"): "captainamerica.png",
    normalize_key("Rhino"): "rhino_new.png",
    normalize_key("M'Baku"): "mbaku_portrait.webp",
    normalize_key("MBaku"): "mbaku_portrait.webp",
    normalize_key("Daredevil (Classic)"): "daredevil.png",
    normalize_key("Daredevil"): "daredevil.png",
    normalize_key("The Maker"): "portrait_maker.png",
    normalize_key("Kang the Conqueror"): "kang.png",
    normalize_key("Kang"): "kang.png",
    normalize_key("Cyclops (Blue Team)"): "Cyclops (Blue Team).png",
    normalize_key("Cyclops (New Xavier School)"): "Cyclops (New Xavier School).png",
    normalize_key("Storm (Pyramid X)"): "stormpyramidx.png",
    normalize_key("Deadpool (X-Force)"): "deadpoolxforce.png",
    normalize_key("Iron Fist (Immortal)"): "Iron Fist (Immortal).png",
    normalize_key("Thanos (Deathless)"): "Thanos (Deathless).png",
    normalize_key("She-Hulk (Deathless)"): "She-Hulk (Deathless).png",
    normalize_key("Guillotine (Deathless)"): "Guillotine (Deathless).webp",
    normalize_key("King Groot (Deathless)"): "kinggrootdeathless.webp",
}

# Clean standard name mapping
CANONICAL_NAMES = {
    normalize_key("The Champion"): "The Champion",
    normalize_key("Okoye"): "Okoye",
    normalize_key("Spider-Man (Classic)"): "Spider-Man (Classic)",
    normalize_key("Spider all"): "Spider-Man (Classic)",
    normalize_key("CGR"): "Cosmic Ghost Rider",
    normalize_key("BLCV"): "Black Widow (Claire Voyant)",
    normalize_key("ABS man"): "Absorbing Man",
    normalize_key("Archangle"): "Archangel",
    normalize_key("Apocalypes"): "Apocalypse",
    normalize_key("Doom"): "Doctor Doom",
    normalize_key("Nicro"): "Nico Minoru",
    normalize_key("Human Tourch"): "Human Torch",
    normalize_key("Nimord"): "Nimrod",
    normalize_key("shuri"): "Shuri",
    normalize_key("Infamous Iron man"): "Iron Man (Infamous)",
    normalize_key("VIV Vision"): "Viv Vision",
    normalize_key("IronHeart"): "Ironheart",
    normalize_key("Hit Mokey"): "Hit-Monkey",
    normalize_key("NickFury"): "Nick Fury",
    normalize_key("Hulking"): "Hulkling",
    normalize_key("Serpent"): "The Serpent",
    normalize_key("Hulk (imortal)"): "Hulk (Immortal)",
    normalize_key("Wepon X"): "Wolverine (Weapon X)",
    normalize_key("Baran Zemo"): "Baron Zemo",
    normalize_key("Hawkey"): "Hawkeye",
    normalize_key("Rocker Raccon"): "Rocket Raccoon",
    normalize_key("VOODOO"): "Doctor Voodoo",
    normalize_key("Cheelith"): "Chee'ilth",
    normalize_key("MR.knight"): "Mr. Knight",
    normalize_key("Jeo Fixit"): "Joe Fixit",
    normalize_key("miles"): "Spider-Man (Miles Morales)",
    normalize_key("G 2099"): "Guillotine 2099",
    normalize_key("LongShot"): "Longshot",
    normalize_key("Pavitra prabakar"): "Spider-Man (Pavitr Prabhakar)",
    normalize_key("Capatain Britain"): "Captain Britain",
    normalize_key("Magneto(white)"): "Magneto (House of X)",
    normalize_key("Magneto (White)"): "Magneto (House of X)",
    normalize_key("gon"): "Ægon",
    normalize_key("Aegon"): "Ægon",
    normalize_key("Count nefaria"): "Count Nefaria",
    normalize_key("Mole man"): "Mole Man",
    normalize_key("Shehulk"): "She-Hulk",
    normalize_key("Anti Venom"): "Anti-Venom",
}

def resolve_champion_image(champ_name: str, champ_class: str = "") -> str:
    norm = normalize_key(champ_name)
    
    # 1. Explicit alias match (High Priority)
    if norm in ALIASES:
        return f"/images/{ALIASES[norm]}"
    
    # 2. Exact match in IMAGE_MAP
    if norm in IMAGE_MAP:
        return f"/images/{IMAGE_MAP[norm]}"
    
    # 3. Class fallback icon
    class_norm = champ_class.lower() if champ_class else "cosmic"
    return f"/images/classes/{class_norm}.svg"

def load_all_mcoc_data() -> Dict[str, Any]:
    wb = openpyxl.load_workbook(DATASET_PATH, data_only=True)

    # 1. Build champion list from available sheets
    # First, collect class info from Upgrade Plan sheet
    class_map: Dict[str, str] = {}
    if "Upgrade Plan" in wb.sheetnames:
        sheet = wb["Upgrade Plan"]
        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(r, 2).value
            c_class = sheet.cell(r, 8).value
            if name and c_class:
                raw_k = normalize_key(str(name).strip())
                canonical_name = CANONICAL_NAMES.get(raw_k, str(name).strip())
                canon_k = normalize_key(canonical_name)
                class_map[canon_k] = str(c_class).strip()

    # Collect all unique champion names from Chamipons_tag and Immunity sheets
    all_champ_names: Dict[str, str] = {}  # normalized_key -> display_name

    for sheet_name in ["Chamipons_tag", "Immunity"]:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for r in range(2, sheet.max_row + 1):
                name = sheet.cell(r, 1).value
                if name and str(name).strip():
                    c_name = str(name).strip()
                    k = normalize_key(c_name)
                    if k not in all_champ_names:
                        all_champ_names[k] = CANONICAL_NAMES.get(k, c_name)

    # Build a comprehensive fallback class map
    fallback_classes = {
        "abominationimmortal": "Science", "adamwarlock": "Cosmic", "agathaharkness": "Mystic", "agentvenom": "Skill", "airwalker": "Cosmic", "angela": "Cosmic", "annihilus": "Cosmic", "apocalypse": "Mutant", "archangel": "Mutant", "attuma": "Skill", "bastion": "Tech", "betaraybill": "Cosmic", "blackbolt": "Cosmic", "blacktarantula": "Skill", "blackwidow": "Skill", "blade2": "Skill", "bladestellarforged": "Skill", "bullseye": "Skill", "cable": "Mutant", "captainamerica": "Science", "captainamericawwii": "Science", "captainamericainfinitywar": "Science", "captainmarvel": "Cosmic", "captainmarvelmovie": "Cosmic", "carnage": "Cosmic", "civilwarrior": "Tech", "corvusglaive": "Cosmic", "cosmicghostrider": "Cosmic", "crossbones": "Skill", "cullobsidian": "Cosmic", "cyclopsnewxavierschool": "Mutant", "danimoonstar": "Mutant", "darkphoenix": "Mutant", "darkhawk": "Tech", "deadpool": "Mutant", "deadpoolxforce": "Mutant", "diablo": "Mystic", "doctorbong": "Science", "doctordoom": "Mystic", "doctoroctopus": "Tech", "doctorstrange": "Mystic", "dormammu": "Mystic", "dracula": "Mystic", "drax": "Cosmic", "dust": "Mutant", "electro": "Science", "falconjoaquintorres": "Skill", "galan": "Cosmic", "gamora": "Cosmic", "gentle": "Mutant", "gladiator": "Cosmic", "goldpool": "Mutant", "gorr": "Cosmic", "groot": "Cosmic", "guardian": "Tech", "guillotine": "Mystic", "guillotinedeathless": "Mystic", "havok": "Mutant", "heimdall": "Cosmic", "hela": "Cosmic", "hercules": "Cosmic", "hobgoblinphilurich": "Tech", "hulkling": "Cosmic", "hyperion": "Cosmic", "ikaris": "Cosmic", "imperiosa": "Mystic", "invisiblewoman": "Science", "ironfist": "Mystic", "ironfistimmortal": "Mystic", "ironpatriot": "Tech", "jackolantern": "Tech", "jeangrey": "Mutant", "jubilee": "Mutant", "kang": "Tech", "kangtheconqueror": "Tech", "karnak": "Skill", "karolinadean": "Cosmic", "kinggroot": "Cosmic", "kinggrootdeathless": "Cosmic", "knull": "Cosmic", "kraven": "Skill", "ladydeathstrike": "Tech", "lizard": "Science", "loki": "Mystic", "madelynepryor": "Mutant", "maestro": "Science", "magik": "Mystic", "manthing": "Mystic", "mangog": "Mystic", "masacre": "Skill", "medusa": "Cosmic", "mistyknight": "Skill", "moonknight": "Skill", "msmarvel": "Cosmic", "mskamalakhan": "Cosmic", "namor": "Mutant", "nebula": "Tech", "nickfury": "Skill", "nightthrasher": "Skill", "nimrod": "Tech", "nova": "Cosmic", "odin": "Cosmic", "okoye": "Skill", "oldmanlogan": "Mutant", "omegared": "Mutant", "patriot": "Science", "phoenix": "Cosmic", "phylavell": "Cosmic", "pixie": "Mutant", "platinumpool": "Mutant", "proximamidnight": "Cosmic", "punisher": "Skill", "purgatory": "Mystic", "quake": "Science", "quicksilver": "Science", "redgoblin": "Cosmic", "rhino": "Science", "rintrah": "Mystic", "ronan": "Cosmic", "ronin": "Skill", "rubythursday": "Science", "sabretooth": "Mutant", "scarletwitch": "Mystic", "scarletwitchclassic": "Mystic", "scarletwitchsigil": "Mystic", "scarletwitchmarvel": "Mystic", "scream": "Cosmic", "sersi": "Cosmic", "shatterstar": "Mutant", "shehulkdeathless": "Science", "shuri": "Tech", "silversurfer": "Cosmic", "solvarch": "Cosmic", "sorcerersupreme": "Mystic", "spiderman2099": "Science", "spidermanmilesmorales": "Science", "spiderham": "Science", "spiderman_pavitr": "Mystic", "spidermansymbiote": "Cosmic", "spiderwoman": "Science", "squirrelgirl": "Skill", "starlord": "Tech", "starlord2": "Tech", "starlordstellarforged": "Tech", "stryfe": "Mutant", "summonedsymbioid": "Cosmic", "sunspot": "Mutant", "superskull": "Cosmic", "superiorironman": "Cosmic", "symbiotesupreme": "Mystic", "terrax": "Cosmic", "thanos": "Cosmic", "thanosdeathless": "Cosmic", "thechampion": "Cosmic", "hood": "Mystic", "thehood": "Mystic", "theleader": "Science", "themaker": "Science", "theoverseer": "Science", "theserpent": "Cosmic", "thor": "Cosmic", "thorragnarok": "Skill", "toad": "Mutant", "ultron": "Tech", "ultronclassic": "Tech", "unstoppablecolossus": "Mystic", "venom": "Cosmic", "venomtheduck": "Cosmic", "venompool": "Cosmic", "visionaarkus": "Cosmic", "vision": "Tech", "visionageofultron": "Tech", "visiondeathless": "Tech", "void": "Science", "vox": "Cosmic", "warmachine": "Tech", "wave": "Science", "wolverine": "Mutant", "wolverineweaponx": "Mutant", "wong": "Mystic", "yelenabelova": "Skill", "yondu": "Tech", "aegon": "Skill"
    }

    # Also extract from Story champions early to catch any others
    if "Story champions" in wb.sheetnames:
        sheet = wb["Story champions"]
        for r_idx, cls_name in enumerate(["Cosmic", "Skill", "Mutant", "Mystic", "Science", "Tech"], start=2):
            for col in range(1, 7):
                val = sheet.cell(r_idx, col).value
                if val:
                    raw_k = normalize_key(str(val).strip())
                    canon_k = normalize_key(CANONICAL_NAMES.get(raw_k, str(val).strip()))
                    fallback_classes[canon_k] = cls_name
        for col_idx, cls_name in enumerate(["Cosmic", "Mystic", "Mutant", "Science", "Skill", "Tech"], start=1):
            for r in range(14, 30):
                val = sheet.cell(r, col_idx).value
                if val and not str(val).strip().lower().startswith(('best', 'support', 'pre-fight', '1 time')):
                    raw_k = normalize_key(str(val).strip())
                    canon_k = normalize_key(CANONICAL_NAMES.get(raw_k, str(val).strip()))
                    fallback_classes[canon_k] = cls_name

    # Build champions_list from all unique names
    champions_list = []
    champ_by_name: Dict[str, Dict[str, Any]] = {}
    for i, (k, c_name) in enumerate(sorted(all_champ_names.items(), key=lambda x: x[1])):
        raw_k = normalize_key(c_name)
        canon_k = normalize_key(CANONICAL_NAMES.get(raw_k, c_name))
        c_cls = class_map.get(canon_k, fallback_classes.get(canon_k, "Cosmic"))
        img_url = resolve_champion_image(c_name, c_cls)
        champ_obj = {
            "id": i + 1,
            "s_no": i + 1,
            "name": c_name,
            "class": c_cls,
            "image": img_url,
            "rarities": [7, 6, 5],
            "rating": 4.9 if c_cls in ["Cosmic", "Mystic", "Mutant"] else 4.8,
            "immunities": [],
            "tags": [],
            "tier": "A-Tier",
            "prestige": 14850 if i % 5 == 0 else 14200 + (i % 500)
        }
        champions_list.append(champ_obj)
        champ_by_name[k] = champ_obj

    # 2. Main sheet - Immunities with descriptions & champion lists
    immunities_catalog = []
    if "Main" in wb.sheetnames:
        sheet = wb["Main"]
        for r in range(2, sheet.max_row + 1):
            imm_name = sheet.cell(r, 1).value
            desc = sheet.cell(r, 2).value
            count = sheet.cell(r, 3).value
            champ_str = sheet.cell(r, 4).value
            if imm_name:
                imm_obj = {
                    "name": str(imm_name).strip(),
                    "description": str(desc).strip() if desc else f"Immunity to {imm_name} debuffs and damage over time.",
                    "count": int(count) if count and str(count).replace('.0','').isdigit() else 0,
                    "champions": [c.strip() for c in str(champ_str).split(",") if c.strip()] if champ_str else []
                }
                immunities_catalog.append(imm_obj)
                
                for c_item in imm_obj["champions"]:
                    c_key = normalize_key(c_item)
                    if c_key in champ_by_name:
                        if imm_obj["name"] not in champ_by_name[c_key]["immunities"]:
                            champ_by_name[c_key]["immunities"].append(imm_obj["name"])

    # 3. Immunity Matrix Sheet
    if "Immunity" in wb.sheetnames:
        sheet = wb["Immunity"]
        headers = [sheet.cell(1, c).value for c in range(2, sheet.max_column + 1)]
        for r in range(2, sheet.max_row + 1):
            c_name = sheet.cell(r, 1).value
            if not c_name:
                continue
            c_key = normalize_key(str(c_name))
            if c_key in champ_by_name:
                for c_idx, h in enumerate(headers):
                    if h and sheet.cell(r, c_idx + 2).value == "Yes":
                        imm_clean = str(h).strip()
                        if imm_clean not in champ_by_name[c_key]["immunities"]:
                            champ_by_name[c_key]["immunities"].append(imm_clean)

    # 4. Tags & Categories (2,959 rows)
    all_tags = set()
    all_categories = set()
    grouped_tags: Dict[str, set] = {}

    for tag_sheet in ["Chamipons_tag", "Story_nodes"]:
        if tag_sheet in wb.sheetnames:
            sheet = wb[tag_sheet]
            for r in range(2, sheet.max_row + 1):
                c_name = sheet.cell(r, 1).value
                tag = sheet.cell(r, 2).value
                cat = sheet.cell(r, 3).value
                
                tag_clean = str(tag).strip() if tag and str(tag).strip() else None
                cat_clean = str(cat).strip() if cat and str(cat).strip() else "General"
                
                if tag_clean:
                    all_tags.add(tag_clean)
                    all_categories.add(cat_clean)
                    if cat_clean not in grouped_tags:
                        grouped_tags[cat_clean] = set()
                    grouped_tags[cat_clean].add(tag_clean)

                if c_name:
                    c_key = normalize_key(str(c_name))
                    if c_key in champ_by_name:
                        if "categories" not in champ_by_name[c_key]:
                            champ_by_name[c_key]["categories"] = []
                        if tag_clean and tag_clean not in champ_by_name[c_key]["tags"]:
                            champ_by_name[c_key]["tags"].append(tag_clean)
                        if cat_clean and cat_clean not in champ_by_name[c_key]["categories"]:
                            champ_by_name[c_key]["categories"].append(cat_clean)

    # 5. Story Champions & Accurate Class Meta Tier Lists
    # Columns in Row 13 of Story champions:
    # Col 1 (A) = Cosmic, Col 2 (B) = Mystic, Col 3 (C) = Mutant, Col 4 (D) = Science, Col 5 (E) = Skill, Col 6 (F) = Tech
    story_tiers: Dict[str, List[Dict[str, Any]]] = {
        "Cosmic": [],
        "Skill": [],
        "Mutant": [],
        "Mystic": [],
        "Science": [],
        "Tech": [],
        "Synergy Support": []
    }

    if "Story champions" in wb.sheetnames:
        sheet = wb["Story champions"]
        
        # Section A: S-Tier Champions (Rows 2 to 7)
        s_tier_row_classes = ["Cosmic", "Skill", "Mutant", "Mystic", "Science", "Tech"]
        for r_idx, cls_name in enumerate(s_tier_row_classes, start=2):
            for col in range(1, 7):
                c_val = sheet.cell(r_idx, col).value
                if c_val and str(c_val).strip():
                    raw_name = str(c_val).strip()
                    k = normalize_key(raw_name)
                    display_name = CANONICAL_NAMES.get(k, raw_name)
                    img = resolve_champion_image(display_name, cls_name)
                    
                    if k in champ_by_name:
                        champ_by_name[k]["tier"] = "S-Tier"
                        
                    # Add to S-Tier if not already present
                    if not any(x["name"].lower() == display_name.lower() for x in story_tiers[cls_name]):
                        story_tiers[cls_name].append({
                            "name": display_name,
                            "image": img,
                            "class": cls_name,
                            "tier": "S-Tier",
                            "note": "Top Story & Meta Pick"
                        })

        # Section B: Best Progression Picks per Class (Rows 14 to 29)
        col_classes = ["Cosmic", "Mystic", "Mutant", "Science", "Skill", "Tech"]
        for col_idx, cls_name in enumerate(col_classes, start=1):
            for r in range(14, 30):
                c_val = sheet.cell(r, col_idx).value
                if c_val and str(c_val).strip():
                    raw_name = str(c_val).strip()
                    # Skip invalid non-champion text
                    if raw_name.lower().startswith(('best ', 'support', 'pre-fight', '1 time')):
                        continue
                    k = normalize_key(raw_name)
                    display_name = CANONICAL_NAMES.get(k, raw_name)
                    img = resolve_champion_image(display_name, cls_name)
                    
                    if not any(x["name"].lower() == display_name.lower() for x in story_tiers[cls_name]):
                        story_tiers[cls_name].append({
                            "name": display_name,
                            "image": img,
                            "class": cls_name,
                            "tier": "A-Tier" if len(story_tiers[cls_name]) >= 5 else "S-Tier",
                            "note": "High-Value Quest Pick"
                        })

        # Section C: Synergy & Pre-Fight Supports (Rows 34 to 37)
        support_picks = [
            {"name": "Heimdall", "class": "Cosmic", "note": "1-time Death Cheat & Fury Buff"},
            {"name": "Hela", "class": "Cosmic", "note": "Indestructible Synergy"},
            {"name": "Odin", "class": "Cosmic", "note": "Pre-Fight Buffs (Aptitude, Protection, Resist)"},
            {"name": "Galan", "class": "Cosmic", "note": "Pre-Fight Planetary Seed"}
        ]
        for sup in support_picks:
            img = resolve_champion_image(sup["name"], sup["class"])
            story_tiers["Synergy Support"].append({
                "name": sup["name"],
                "image": img,
                "class": sup["class"],
                "tier": "Support MVP",
                "note": sup["note"]
            })

    # 6. Upgrade Plan Presets (218 items)
    upgrade_plan_presets = []
    if "Upgrade Plan" in wb.sheetnames:
        sheet = wb["Upgrade Plan"]
        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(r, 2).value
            rarity = sheet.cell(r, 3).value
            awaken = sheet.cell(r, 4).value
            rank = sheet.cell(r, 5).value
            future = sheet.cell(r, 6).value
            c_class = sheet.cell(r, 8).value
            importance = sheet.cell(r, 10).value
            priority = sheet.cell(r, 9).value or sheet.cell(r, 11).value
            
            if name:
                c_name = str(name).strip()
                k = normalize_key(c_name)
                display_name = CANONICAL_NAMES.get(k, c_name)
                c_cls = str(c_class).strip() if c_class else "Cosmic"
                upg_item = {
                    "id": len(upgrade_plan_presets) + 1,
                    "champion_name": display_name,
                    "class": c_cls,
                    "image": resolve_champion_image(display_name, c_cls),
                    "rarity": int(rarity) if rarity and str(rarity).replace('.0','').isdigit() else 7,
                    "awakened": str(awaken).upper() == "Y",
                    "current_rank": int(rank) if rank and str(rank).replace('.0','').isdigit() else 1,
                    "future_rank": int(future) if future and str(future).replace('.0','').isdigit() else 3,
                    "importance": str(importance).strip() if importance and importance != 'importance' else "Recommended Rankup",
                    "priority": int(priority) if priority and str(priority).replace('.0','').isdigit() else 2
                }
                upgrade_plan_presets.append(upg_item)

    # 7. Glossary
    glossary = []
    if "Sheet8" in wb.sheetnames:
        sheet = wb["Sheet8"]
        for r in range(4, sheet.max_row + 1):
            term = sheet.cell(r, 1).value
            desc = sheet.cell(r, 2).value
            tips = sheet.cell(r, 3).value
            if term and desc:
                glossary.append({
                    "term": str(term).strip(),
                    "description": str(desc).strip(),
                    "tips": str(tips).strip() if tips else ""
                })

    # 8. Duel Targets
    duel_targets = [
        {"champion": "Hercules", "class": "Cosmic", "player": "DorkLessons", "duel_code": "HERC_DUEL", "stars": 7, "image": resolve_champion_image("Hercules", "Cosmic")},
        {"champion": "The Serpent", "class": "Cosmic", "player": "KT1_Gaming", "duel_code": "SERPENT_PRACTICE", "stars": 7, "image": resolve_champion_image("The Serpent", "Cosmic")},
        {"champion": "Onslaught", "class": "Mutant", "player": "Vega_MCOC", "duel_code": "ONSLAUGHT_DEF", "stars": 7, "image": resolve_champion_image("Onslaught", "Mutant")},
        {"champion": "Doctor Doom", "class": "Mystic", "player": "Lagacy", "duel_code": "DOOM_TARGET", "stars": 6, "image": resolve_champion_image("Doctor Doom", "Mystic")},
        {"champion": "Kushala", "class": "Mystic", "player": "Seatin_MCOC", "duel_code": "KUSHALA_7R3", "stars": 7, "image": resolve_champion_image("Kushala", "Mystic")},
        {"champion": "Kate Bishop", "class": "Skill", "player": "Karatemike", "duel_code": "KBISHOP_WAR", "stars": 7, "image": resolve_champion_image("Kate Bishop", "Skill")},
        {"champion": "Bullseye", "class": "Skill", "player": "BeroMan", "duel_code": "BULLSEYE_PVP", "stars": 7, "image": resolve_champion_image("Bullseye", "Skill")},
        {"champion": "Photon", "class": "Science", "player": "Swedeah", "duel_code": "PHOTON_WAR", "stars": 7, "image": resolve_champion_image("Photon", "Science")},
        {"champion": "Human Torch", "class": "Science", "player": "Ilacros", "duel_code": "HTORCH_PRACTICE", "stars": 6, "image": resolve_champion_image("Human Torch", "Science")},
        {"champion": "Nimrod", "class": "Tech", "player": "ProfHoff", "duel_code": "NIMROD_BG", "stars": 7, "image": resolve_champion_image("Nimrod", "Tech")},
        {"champion": "Shuri", "class": "Tech", "player": "BG_Master", "duel_code": "SHURI_7R3", "stars": 7, "image": resolve_champion_image("Shuri", "Tech")}
    ]

    return {
        "champions": champions_list,
        "immunities": immunities_catalog,
        "story_tiers": story_tiers,
        "upgrade_plan_presets": upgrade_plan_presets,
        "glossary": glossary,
        "duel_targets": duel_targets,
        "all_tags": sorted(list(all_tags)),
        "all_categories": sorted(list(all_categories)),
        "tags_by_category": {cat: sorted(list(t_list)) for cat, t_list in sorted(grouped_tags.items())}
    }
