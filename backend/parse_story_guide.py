import openpyxl
import json
import os
import re

def parse_act8_guide(excel_path, output_json_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'Act 8 Guide' not in wb.sheetnames:
        print("Act 8 Guide sheet not found")
        return

    sheet = wb['Act 8 Guide']
    
    act_data = {
        "title": "Act 8",
        "description": "Easy path for competition",
        "quests": []
    }
    
    current_quest = None
    state = "NONE" # NONE, QUEST_NODES, PATH_DEFENDERS, BOSS_NODES, BOSS_PHASES
    
    current_boss_phase = None
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(row):
            continue
            
        # Clean up row (remove trailing Nones)
        row_str = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # Check for Quest Start
        # Format: (None, None, '8.1.1', 'Node Path', None, None, 'Refer Video: ...')
        if row_str[2] and row_str[2].startswith("8.") and not row_str[1]:
            if current_quest:
                act_data["quests"].append(current_quest)
            
            video_link = ""
            for cell in row_str:
                if "Refer Video:" in cell:
                    video_link = cell.replace("Refer Video:", "").strip()
            
            current_quest = {
                "id": row_str[2],
                "video_url": video_link,
                "path_nodes": [],
                "path_defenders": [],
                "boss": {
                    "name": "",
                    "nodes": [],
                    "phases": []
                }
            }
            state = "QUEST_NODES"
            continue
            
        if not current_quest:
            continue
            
        # State Transitions
        if "Path Defenders" in row_str:
            state = "PATH_DEFENDERS"
            continue
        elif "Boss" in row_str:
            state = "BOSS_NODES"
            boss_name = ""
            for i, val in enumerate(row_str):
                if val == "Boss":
                    if i + 2 < len(row_str) and row_str[i+2]:
                        boss_name = row_str[i+2]
            current_quest["boss"]["name"] = boss_name
            continue
        elif "Phase" in row_str[3] or "Phase" in row_str[2] or "Phase" in row_str[1]:
            state = "BOSS_PHASES"
            phase_name = ""
            for val in row_str:
                if "Phase" in val:
                    phase_name = val
                    break
            current_boss_phase = {
                "name": phase_name,
                "steps": []
            }
            current_quest["boss"]["phases"].append(current_boss_phase)
            continue
            
        # Ignore header rows
        if "S.no" in row_str or "S.NO" in row_str or "Step" in row_str or "Steps" in row_str or "Ability" in row_str:
            continue
            
        # Data Extraction based on State
        if state == "QUEST_NODES":
            # (None, None, None, 1.0, 'Defective Defense', 'Effect...', None)
            if row_str[3] and row_str[4] and row_str[5]:
                try:
                    float(row_str[3]) # check if it's a number
                    current_quest["path_nodes"].append({
                        "name": row_str[4],
                        "effect": row_str[5]
                    })
                except ValueError:
                    pass
                    
        elif state == "PATH_DEFENDERS":
            # (None, None, None, 1.0, 'Easy', 'Green Goblin', None)
            if row_str[3] and row_str[4] and row_str[5]:
                try:
                    float(row_str[3])
                    current_quest["path_defenders"].append({
                        "difficulty": row_str[4],
                        "champion": row_str[5]
                    })
                except ValueError:
                    pass
                    
        elif state == "BOSS_NODES":
            # (None, None, None, 1.0, 'Node', 'Effect', None)
            if row_str[3] and row_str[4] and row_str[5]:
                try:
                    float(row_str[3])
                    current_quest["boss"]["nodes"].append({
                        "name": row_str[4],
                        "effect": row_str[5]
                    })
                except ValueError:
                    pass
            # or (None, None, None, 'Ability', 'What it does', 'What YOU should do', None) -> Phase 0 / General Boss Mechanics
            elif not row_str[3] and row_str[4] and row_str[5] and "Corrupted Armor" in row_str[3]:
                # Special cases where boss mechanics are listed before phases
                 current_quest["boss"]["nodes"].append({
                        "name": row_str[3],
                        "effect": row_str[4] + " | Strategy: " + row_str[5]
                 })
                 
            elif not row_str[3] and row_str[4] and row_str[5] and "Trap" in row_str[3]:
                 current_quest["boss"]["nodes"].append({
                        "name": row_str[3],
                        "effect": row_str[4] + " | Strategy: " + row_str[5]
                 })
                 
            # Bahamet / Scytalis general mechanics are often in cols 3, 4, 5
            elif row_str[3] and row_str[4] and row_str[5] and not re.match(r'^\d', row_str[3]):
                current_quest["boss"]["nodes"].append({
                    "name": row_str[3],
                    "effect": f"{row_str[4]} (Counter: {row_str[5]})" if len(row_str) > 5 and row_str[5] else row_str[4]
                })

        elif state == "BOSS_PHASES":
            if current_boss_phase and row_str[3] and row_str[4] and row_str[5]:
                try:
                    float(row_str[3])
                    current_boss_phase["steps"].append({
                        "mechanic": row_str[4],
                        "action": row_str[5]
                    })
                except ValueError:
                    pass

    if current_quest:
        act_data["quests"].append(current_quest)
        
    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(act_data, f, indent=2)
    print(f"Saved story guide data to {output_json_path}")

if __name__ == "__main__":
    parse_act8_guide(
        "../../nexus-frontend/public/excle/MCOC_dataset.xlsx", 
        "../assest/data/act8_guide.json"
    )
